"""Loads the historical exam-question workbook into WP-002 domain models.

Row-level structural validation is deliberately delegated to
``HistoricalStyleReference`` itself (established in WP-002) rather than
re-implemented here, so ingestion cannot silently diverge from the domain
contract.
"""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError

from exam_generator.config.loader import find_project_root, load_app_config
from exam_generator.historical.errors import (
    HistoricalQuestionRowError,
    WorkbookFormatError,
    WorkbookNotFoundError,
    WorkbookSchemaError,
)
from exam_generator.models import HistoricalStyleReference

# V1 workbook contract (see docs/ARCHITECTURE.md). Extra columns
# (categories_json, accuracy_values, distinction_values, source,
# uploaded_at, created_at, updated_at, ...) are ignored by this subsystem.
REQUIRED_HEADERS: tuple[str, ...] = (
    "id",
    "category",
    "question",
    "answer1",
    "answer2",
    "answer3",
    "answer4",
    "correct_answer_id",
)

DEFAULT_WORKBOOK_FILENAME = "questions_full_export.xlsx"


def default_workbook_path() -> Path:
    """The configured default location of the historical workbook.

    Uses the WP-001 project-relative ``data`` directory configuration; the
    filename itself is currently fixed rather than separately configured,
    since there is only ever one historical workbook in V1.
    """
    app_config = load_app_config()
    return find_project_root() / app_config.paths.data_dir / DEFAULT_WORKBOOK_FILENAME


def _select_worksheet(workbook, sheet_name: str | None):
    if sheet_name is not None:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookFormatError(
                f"Worksheet '{sheet_name}' not found; available worksheets: {workbook.sheetnames}"
            )
        return workbook[sheet_name]

    if len(workbook.sheetnames) == 1:
        return workbook[workbook.sheetnames[0]]

    raise WorkbookFormatError(
        f"Workbook has multiple worksheets {workbook.sheetnames}; "
        "pass an explicit sheet_name rather than guessing which one to ingest."
    )


def _build_column_index(header_row: tuple) -> dict[str, int]:
    header_names = [cell for cell in header_row if isinstance(cell, str) and cell.strip()]
    counts = Counter(header_names)

    duplicated_required = sorted(name for name in REQUIRED_HEADERS if counts.get(name, 0) > 1)
    if duplicated_required:
        raise WorkbookSchemaError(
            f"Duplicate required workbook column(s): {', '.join(duplicated_required)}"
        )

    index: dict[str, int] = {}
    for column_number, cell in enumerate(header_row):
        if isinstance(cell, str) and cell.strip() and cell not in index:
            index[cell] = column_number

    missing = [name for name in REQUIRED_HEADERS if name not in index]
    if missing:
        raise WorkbookSchemaError(f"Missing required workbook column(s): {', '.join(missing)}")

    return index


def _is_blank_row(row: tuple) -> bool:
    return all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row)


def load_historical_questions(
    path: Path | str, *, sheet_name: str | None = None
) -> tuple[list[HistoricalStyleReference], list[str]]:
    """Load and validate the historical workbook.

    Returns ``(questions, canonical_categories)`` where ``canonical_categories``
    preserves first-seen workbook order. Raises a
    ``exam_generator.historical.errors.HistoricalIngestionError`` subclass on
    any expected ingestion failure.
    """
    workbook_path = Path(path)

    if not workbook_path.exists():
        raise WorkbookNotFoundError(f"Historical workbook not found: {workbook_path}")
    if not workbook_path.is_file():
        raise WorkbookNotFoundError(f"Historical workbook path is not a file: {workbook_path}")

    try:
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError) as exc:
        raise WorkbookFormatError(f"Could not open historical workbook {workbook_path}: {exc}") from exc

    worksheet = _select_worksheet(workbook, sheet_name)
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise WorkbookSchemaError(
            f"Historical workbook worksheet '{worksheet.title}' has no header row"
        ) from exc

    column_index = _build_column_index(header_row)

    questions: list[HistoricalStyleReference] = []
    categories_order: list[str] = []
    first_row_by_id: dict[int, int] = {}

    row_number = 1  # the header occupies worksheet row 1
    for row in rows:
        row_number += 1

        if _is_blank_row(row):
            continue

        try:
            reference = HistoricalStyleReference(
                historical_question_id=row[column_index["id"]],
                category=row[column_index["category"]],
                question=row[column_index["question"]],
                answers=[
                    row[column_index["answer1"]],
                    row[column_index["answer2"]],
                    row[column_index["answer3"]],
                    row[column_index["answer4"]],
                ],
                correct_answer=row[column_index["correct_answer_id"]],
            )
        except ValidationError as exc:
            raise HistoricalQuestionRowError(
                f"Invalid historical question at worksheet row {row_number}: {exc}"
            ) from exc

        if reference.historical_question_id in first_row_by_id:
            raise HistoricalQuestionRowError(
                f"Duplicate historical question id {reference.historical_question_id} "
                f"at worksheet row {row_number} "
                f"(first seen at worksheet row {first_row_by_id[reference.historical_question_id]})"
            )
        first_row_by_id[reference.historical_question_id] = row_number

        if reference.category not in categories_order:
            categories_order.append(reference.category)

        questions.append(reference)

    if not questions:
        raise WorkbookSchemaError(
            f"Historical workbook worksheet '{worksheet.title}' contains no valid question rows"
        )

    return questions, categories_order
